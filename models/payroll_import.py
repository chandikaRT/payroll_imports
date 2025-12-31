# -*- coding: utf-8 -*-
import base64
import io
import calendar
from datetime import date
from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class HrPayrollImport(models.Model):
    _name = 'hr.payroll.import'
    _description = 'Payroll Import Sheet'

    # ------------------------------------------------------------------
    # header fields
    # ------------------------------------------------------------------
    name = fields.Char(string='Reference', copy=False, index=True)
    month = fields.Selection([
        ('1', 'January'), ('2', 'February'), ('3', 'March'),
        ('4', 'April'), ('5', 'May'), ('6', 'June'),
        ('7', 'July'), ('8', 'August'), ('9', 'September'),
        ('10', 'October'), ('11', 'November'), ('12', 'December')
    ], required=True)
    year = fields.Integer(required=True, default=lambda _: date.today().year)
    line_ids = fields.One2many('hr.payroll.import.line', 'import_id', string='Import Lines')
    state = fields.Selection([('draft', 'Draft'), ('done', 'Done')], default='draft')

    import_file = fields.Binary(string='Excel File')
    import_filename = fields.Char()

    # ------------------------------------------------------------------
    # sequence for name 001/10/2025
    # ------------------------------------------------------------------
    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', 'New') == 'New':
                month = int(vals.get('month'))
                year = int(vals.get('year'))
                seq = self.env['ir.sequence'].next_by_code('hr.payroll.import') or '001'
                vals['name'] = f"{seq}/{month:02d}/{year}"
        return super().create(vals_list)

    # ------------------------------------------------------------------
    # import Excel button
    # ------------------------------------------------------------------
    def action_import_excel(self):
        self.ensure_one()
        if not self.import_file:
            raise UserError(_('Please choose an Excel file first.'))
        if not load_workbook:
            raise UserError(_('python library "openpyxl" is missing on the server.'))

        data = base64.b64decode(self.import_file)
        ws = load_workbook(io.BytesIO(data), data_only=True).active

        error_lines, created = [], 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            row = list(row) + [None] * 5
            emp_code, emp_name, input_type_code, amount, notes = row[:5]

            # Validate numeric amount
            try:
                amount = float(amount or 0)
            except Exception:
                error_lines.append((idx, 'Invalid amount value'))
                continue

            # Find employee
            employee = None
            if emp_code:
                employee = self.env['hr.employee'].search([('barcode', '=', str(emp_code))], limit=1)
            if not employee and emp_name:
                employee = self.env['hr.employee'].search([('name', 'ilike', str(emp_name))], limit=1)
            if not employee:
                error_lines.append((idx, f'Employee not found: {emp_code}/{emp_name}'))
                continue

            # Find input type (must exist beforehand)
            if not input_type_code:
                error_lines.append((idx, 'Input Type Code is required'))
                continue
                
            input_type = self.env['hr.payslip.input.type'].search([
                ('code', '=', str(input_type_code).strip())
            ], limit=1)
            if not input_type:
                error_lines.append((idx, f'Input Type not found: {input_type_code}. Create it in Payroll > Configuration > Input Types.'))
                continue

            # Build description
            desc_parts = [input_type.name]
            if notes:
                desc_parts.append(str(notes))
            description = ' - '.join(desc_parts)

            # Create line
            self.env['hr.payroll.import.line'].create({
                'import_id': self.id,
                'employee_id': employee.id,
                'input_type_id': input_type.id,
                'amount': amount,
                'description': description,
            })
            created += 1

        self.import_file = False
        message = f'Imported {created} rows.'
        if error_lines:
            message += '\nErrors:\n' + '\n'.join([f'Row {r}: {m}' for r, m in error_lines])
        
        # Show result in a non-blocking way
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Result'),
                'message': message,
                'type': 'info' if error_lines else 'success',
                'sticky': False,
                'next': {
                    'type': 'ir.actions.act_window',
                    'res_model': self._name,
                    'res_id': self.id,
                    'view_mode': 'form',
                    'target': 'main',
                }
            }
        }

    # ------------------------------------------------------------------
    # apply to payslips button
    # ------------------------------------------------------------------
    def action_apply_to_payslips(self):
        self.ensure_one()
        Payroll = self.env['hr.payslip']
        Input = self.env['hr.payslip.input']

        for line in self.line_ids:
            year = int(self.year)
            month = int(self.month)
            last_day = calendar.monthrange(year, month)[1]
            date_from = date(year, month, 1)
            date_to = date(year, month, last_day)

            # Find or create payslip
            slip = Payroll.search([
                ('employee_id', '=', line.employee_id.id),
                ('date_from', '<=', date_to),
                ('date_to', '>=', date_from),
            ], limit=1)

            if not slip:
                contract = self.env['hr.contract'].search([
                    ('employee_id', '=', line.employee_id.id),
                    ('state', 'in', ('open', 'close'))
                ], limit=1)
                if not contract:
                    continue

                struct = (contract.structure_type_id.default_struct_id or
                          self.env['hr.payroll.structure'].search([], limit=1))
                if not struct:
                    raise UserError(_('No salary structure for employee %s') % line.employee_id.name)
                
                slip = Payroll.create({
                    'employee_id': contract.employee_id.id,
                    'contract_id': contract.id,
                    'struct_id': struct.id,
                    'date_from': date_from,
                    'date_to': date_to,
                })

            # Update or create payslip input
            existing = Input.search([
                ('payslip_id', '=', slip.id),
                ('input_type_id', '=', line.input_type_id.id)
            ], limit=1)
            
            if existing:
                existing.write({
                    'amount': line.amount,
                    'name': line.description or '',
                })
            else:
                Input.create({
                    'payslip_id': slip.id,
                    'input_type_id': line.input_type_id.id,
                    'amount': line.amount,
                    'name': line.description or '',
                })

            line.applied = True

        self.state = 'done'
        return {'type': 'ir.actions.act_window_close'}


class HrPayrollImportLine(models.Model):
    _name = 'hr.payroll.import.line'
    _description = 'Payroll Import Line'

    import_id = fields.Many2one('hr.payroll.import', ondelete='cascade')
    employee_id = fields.Many2one('hr.employee', string='Employee', required=True)
    input_type_id = fields.Many2one('hr.payslip.input.type', string='Input Type', required=True)
    amount = fields.Monetary(string='Amount', currency_field='company_currency', required=True)
    description = fields.Char(string='Description')
    applied = fields.Boolean(default=False)

    company_currency = fields.Many2one('res.currency', related='employee_id.company_id.currency_id', readonly=True)