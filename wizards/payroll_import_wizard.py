import base64
import io
import calendar
from datetime import date
from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class PayrollImportWizard(models.TransientModel):
    _name = 'hr.payroll.import.wizard'
    _description = 'Import Payroll from Excel'

    file = fields.Binary('Excel File', required=True)
    filename = fields.Char('File Name')
    month = fields.Selection([(str(i), str(i)) for i in range(1, 13)], string='Month', required=True)
    year = fields.Integer('Year', required=True)
    import_ref = fields.Many2one('hr.payroll.import', string='Import Sheet')

    # ------------------------------------------------------------------
    # import button
    # ------------------------------------------------------------------
    def action_import(self):
        if not load_workbook:
            raise UserError(_('openpyxl is not installed. Run: pip install openpyxl'))
        if not self.file:
            raise UserError(_('Please upload a file.'))

        data = base64.b64decode(self.file)
        ws = load_workbook(io.BytesIO(data), data_only=True).active

        # Create or reuse import sheet
        sheet = self.import_ref or self.env['hr.payroll.import'].create({
            'name': f'IMPORT/{self.year}/{self.month}',
            'month': self.month,
            'year': self.year
        })

        error_lines, created = [], 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            row = list(row) + [None] * 5
            emp_code, emp_name, input_type_code, amount, notes = row[:5]

            # Validate amount
            try:
                amount = float(amount or 0)
            except Exception:
                error_lines.append((idx, 'Invalid amount value'))
                continue

            # Find employee
            employee = None
            if emp_code:
                employee = self.env['hr.employee'].search([('barcode', '=', str(emp_code))], limit=1)
            if not employee and emp_name:
                employee = self.env['hr.employee'].search([('name', 'ilike', str(emp_name))], limit=1)
            if not employee:
                error_lines.append((idx, f'Employee not found: {emp_code}/{emp_name}'))
                continue

            # Find input type
            if not input_type_code:
                error_lines.append((idx, 'Input Type Code is required'))
                continue
                
            input_type = self.env['hr.payslip.input.type'].search([
                ('code', '=', str(input_type_code).strip())
            ], limit=1)
            if not input_type:
                error_lines.append((idx, f'Input Type not found: {input_type_code}'))
                continue

            # Build description
            desc_parts = [input_type.name]
            if notes:
                desc_parts.append(str(notes))
            description = ' - '.join(desc_parts)

            # Create line
            self.env['hr.payroll.import.line'].create({
                'import_id': sheet.id,
                'employee_id': employee.id,
                'input_type_id': input_type.id,
                'amount': amount,
                'description': description,
            })
            created += 1

        # Clear file
        self.file = False

        # Return to the import sheet
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'hr.payroll.import',
            'res_id': sheet.id,
            'view_mode': 'form',
            'target': 'main',
        }